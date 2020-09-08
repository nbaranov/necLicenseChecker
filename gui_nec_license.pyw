#!/usr/bin/python3

import openpyxl
import requests
import PySimpleGUI as sg
import threading
from bs4 import BeautifulSoup as bs


# read logins and password from first sheet xlsX file
def readLogPas(path):
    wb = openpyxl.open(path)
    sheets = wb.sheetnames
    for sh in sheets:
        if sh[0:15] == "Логины и пароли":
            sheet = wb[sh]

    logpass = []
    logins = [v[0].value for v in sheet['A2:A1000']]
    passwords = [v[0].value for v in sheet['B2:B1000']]
    for i in range(len(logins)-1, -1, -1):
        if logins[i] == None or passwords[i] == None:
            logins.pop(i)
            passwords.pop(i)
        else:
            logpass.append({
                "login" : logins[i].strip(" .,"),
                "password" : passwords[i].strip(" .,"),
                "status" : ""
            })

    return logpass
readLogPas("nec_license.xlsx")

def writeTables(path, license, logins):
    # date for name sheet 
    import datetime
    date = datetime.date.today()

    # open workbook and create sheet
    wb = openpyxl.open(path)  
    sheet = wb.create_sheet(f"Логины и пароли {date}")
    # write table
    i = 1
    sheet.cell(i, 1, 'Login')
    sheet.cell(i, 2, 'Password')
    sheet.cell(i, 3, 'Status')

    for line in logins:
        i += 1
        sheet.cell(i, 1, line["login"])
        sheet.cell(i, 2, line["password"])
        sheet.cell(i, 3, line['status'])

    sheet = wb.create_sheet(f"Остатки лицензий {date}")
    # write table
    i = 1
    sheet.cell(i, 1, 'Project Name')
    sheet.cell(i, 2, 'P/O Number')
    sheet.cell(i, 3, 'Product Type')
    sheet.cell(i, 4, 'Key Type')
    sheet.cell(i, 5, 'Parameter')
    sheet.cell(i, 6, 'Unused')
    sheet.cell(i, 7, "Login")
    sheet.cell(i, 8, "Password")
    for line in license:
        i += 1
        sheet.cell(i, 1, line['Project Name'])
        sheet.cell(i, 2, line['PO Number'])
        sheet.cell(i, 3, line['Product Type'])
        sheet.cell(i, 4, line['Key Type'])
        sheet.cell(i, 5, line['Parameter'])
        sheet.cell(i, 6, line['Unusued'])
        sheet.cell(i, 7, line["Login"])
        sheet.cell(i, 8, line["Password"])
    try:
        wb.save(path)
        print(f"Логины и лицензии сохранены в файл \n\t{path}")
    except:
        wb.save(f"nec_license {date}.xlsx")
        print(f"Логины и лицензии сохранены в файл \n\t nec_license {date}.xlsx")

def checkLicense(path, window):

    headers = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) snap Chromium/83.0.4103.61 Chrome/83.0.4103.61 Safari/537.36"
    }

    license = []
    logins = readLogPas(path=path)

    countlogins = 0
    lenlogins = len(logins)
    print(f"Собранно {lenlogins} логинов. Начинаю проверку")
    window.refresh()
    for logpas in logins:
        countlogins += 1
        log = logpas["login"]
        print(f"Проверяю {log} {countlogins} из {lenlogins}")
        window.refresh()     
        session = requests.Session()
        session.headers.update(headers)
        session.get('https://www.nec-pasolink-softwarekey.com/tws/twsas/twsas011.aspx')
        request = session.get('https://www.nec-pasolink-softwarekey.com/tws/TWSAS/TWSAS021.aspx')

        # read __VIEWSTATE and __VIEWSTATEGENERATOR from page
        soup = bs(request.text, "lxml")
        inputs = soup.findAll(name="input")
        viewstate = inputs[0].attrs["value"]
        viewstategenegator = inputs[1].attrs["value"]

        # create data for auth
        data = {
            "__VIEWSTATE": viewstate,
            "__VIEWSTATEGENERATOR": viewstategenegator,
            "txt_USERID": logpas["login"],
            "txt_PSW": logpas["password"],
            "btn_Login": "Login",
            "hdn_ErrorMessage": "",
            "hdn_ErrorName": ""
        }

        # auth with data
        request = session.post("https://www.nec-pasolink-softwarekey.com/tws/TWSAS/TWSAS021.aspx", data=data)
        try:
            # check status auth
            soup = bs(request.text, "lxml")
            msg = soup.find(name="span", attrs={"id": "lbl_message"}).text
            #print(msg)
            if msg[0:25] == "You have failed to login.":
                print("Не существует")
                window.refresh()
                logpas["status"] = "Не существует"
                continue
            elif msg[0:64] == "Your account has been locked because of inactivity for 180 days.":
                print("Заблокирован")
                logpas["status"] = "Заблокирован"
                continue
            elif msg[0:28] == "Only alphanumeric characters":
                print("Недопустимый символ в логине")
                logpas["status"] = "Недопустимый символ в логине"
                continue
        except:
            # read new __VIEWSTATE __VIEWSTATEGENERATOR    
            request = session.get("https://www.nec-pasolink-softwarekey.com/tws/TWSBS/TWSBS011.aspx")
            soup = bs(request.text, "lxml")
            viewstate = soup.find(name="input", attrs={"id" : "__VIEWSTATE"}).attrs["value"]
            viewstategenegator = soup.find(name="input", attrs={"id" : "__VIEWSTATEGENERATOR"}).attrs["value"]

            # create new data
            data= {
                "__EVENTTARGET": "lnkButton_60_0",
                "__EVENTARGUMENT": "",
                "__VIEWSTATE": viewstate,
                "__VIEWSTATEGENERATOR": viewstategenegator,      
            } 

            # get license information
            request = session.post("https://www.nec-pasolink-softwarekey.com/tws/TWSBS/TWSBS011.aspx", data=data)    
            soup = bs(request.text, "lxml")
            siteTable = soup.find("table", attrs={"class": "DataTable"})
            siteTable = siteTable.findAll("tr")

                #create table without rowspans
            try:      
                rowspan = int(siteTable[1].find("td").get("rowspan"))
                table = []
                for i in range(rowspan):
                    table.append([None,None,None,None,None,None,None,None,None,None,])
                for i in range(1, len(siteTable)):
                    row = siteTable[i].findAll("td")
                    for j in range(len(row)):
                        start = j       
                        try:
                            rowspan = int(row[j].get("rowspan"))
                            for k in range(rowspan):
                                if i > 1:
                                    start = j+4
                                for l in range(start,10):
                                    if table[i+k-1][l] == None:
                                        table[i+k-1][l] = row[j].text
                                        break
                                    else:
                                        continue
                                                                            
                        
                        except:
                            if i > 1:
                                start = j+4
                            for l in range(start,10):
                                if table[i-1][l] == None:
                                    table[i-1][l] = row[j].text
                                    break
                                else:
                                    continue

            # apend license list

                count = 0
                for row in table:
                    if row[9] != "0":
                        count += 1
                        license.append({
                            "Project Name" : row[2],
                            "PO Number" : row[3],
                            "Product Type" : row[4],
                            "Key Type" : row[5],
                            "Parameter" : row[6],
                            "Unusued" : row[9],
                            "Login" : logpas["login"],
                            "Password" : logpas["password"]
                        })
                logpas["status"] = f"Содержит лицензи"
                print(f"Содержит лицензий: {count}")
            except:
                logpas["status"] = "Израсходован полностью"
                print("Израсходован полностью")
            print("Вылогиниваюсь")
            window.refresh()

            # logout from site

            request = session.get("https://www.nec-pasolink-softwarekey.com/tws/TWSES/TWSES011.aspx")
            soup = bs(request.text, "lxml")
            viewstate = soup.find(name="input", attrs={"id" : "__VIEWSTATE"}).attrs["value"]
            viewstategenegator = soup.find(name="input", attrs={"id" : "__VIEWSTATEGENERATOR"}).attrs["value"]

            # create new data
            data= {
                "__EVENTTARGET": "",
                "__EVENTARGUMENT": "",
                "__VIEWSTATE": viewstate,
                "__VIEWSTATEGENERATOR": viewstategenegator,
                "hdn_ErrorMessage": "", 
                "btn_Menu": "Menu",
                "hdn_Sort": "31"
            }

            request = session.post("https://www.nec-pasolink-softwarekey.com/tws/TWSES/TWSES011.aspx", data=data) 

            request = session.get("https://www.nec-pasolink-softwarekey.com/tws/TWSBS/TWSBS011.aspx")
            soup = bs(request.text, "lxml")
            viewstate = soup.find(name="input", attrs={"id" : "__VIEWSTATE"}).attrs["value"]
            viewstategenegator = soup.find(name="input", attrs={"id" : "__VIEWSTATEGENERATOR"}).attrs["value"]
            # create new data
            data= {
                "__EVENTTARGET": "lnk_LogOut",
                "__EVENTARGUMENT": "",
                "__VIEWSTATE": viewstate,
                "__VIEWSTATEGENERATOR": viewstategenegator,
            }

            session.post("https://www.nec-pasolink-softwarekey.com/tws/TWSBS/TWSBS011.aspx", data=data)
            session.close()

    writeTables(path=path, license=license, logins=logins)




# All the stuff inside your window.
layout = [  [sg.Text('Выберете файл .xlsx, в котором есть лист с именем "Логины и пароли"')],
            [sg.Text('На котором ячейки A2:А1000 - логины, B2:В1000 - пароли.')],
            [sg.Input("nec_license.xlsx", size=(57, 10), key="path"), sg.FileBrowse('Выбрать файл')],
            [sg.Button('Проверить лицензии', size=(63,1))],
            [sg.Output(size=(70,15),background_color="black", text_color=("white"))]
        ]

# Create the Window
window = sg.Window('NEC License Сhecker by nikita.baranov@nokia.com', layout)

# Event Loop to process "events" and get the "values" of the inputs
while True:
    event, values = window.read(timeout=100)
    if event == sg.WIN_CLOSED or None:	# if user closes window or clicks cancel
        break
    elif event == "Проверить лицензии":
        thread = threading.Thread(target=checkLicense, args=(values["path"], window), daemon=True)
        thread.start()

window.close()
